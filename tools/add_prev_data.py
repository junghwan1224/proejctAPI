from openpyxl import load_workbook
import os
import sys
from pprint import pprint
import requests

# API URL 설정:
BASE_PATH = 'http://localhost:3001/admin{}'

# ACCESS_TOKEN을 위한 STAFF 로그인 정보:
LOGIN_DATA = {
    'email': 'chianti@kakao.com',
    'password': '1234'
}

# 새 STAFF 생성 여부:
CREATE_STAFF = False

# 새로 생성할 STAFF 데이터:
CREATE_STAFF_DATA = [
    {'email': 'chianti@kakao.com',
     'password': '1234',
     'name': '정구현',
     'phone': '01024733891',
     'department': '개발팀',
     'rank': '팀원',
     'permission': 'READ_PRODUCT'
     }
]

# 엑셀 데이터 저장 위치:
PURCHASE_BOOK_PATH = '~/Documents/ujshaft-resources/매입 내역.xlsx'
SALES_BOOK_PATH = '~/Documents/ujshaft-resources/매출 내역.xlsx'
CLIENT_BOOK_PATH = '~/Documents/ujshaft-resources/거래처 관리.xlsx'
PRODUCT_BOOK_PATH = '~/Documents/ujshaft-resources/품목 관리.xlsx'


class Cells:
    def __init__(self, worksheet):
        self._worksheet = worksheet
        self.max_row = worksheet.max_row

    def get(self, col_name, row_num):
        value = self._worksheet['{}{}'.format(col_name, row_num)].value
        return '' if value is None else value


def main():
    # 필요한 경우 유저 생성:
    if CREATE_STAFF:
        for staff in CREATE_STAFF_DATA:
            req = requests.post(BASE_PATH.format('/staff'), staff)
            if req.status_code != 201:
                print('[*] Unable to create new staff data.',
                      "Please check API's permission.")
                sys.exit()

    # 로그인 후 AccessToken 취득:
    req = requests.post(BASE_PATH.format('/login'), LOGIN_DATA)
    if req.status_code != 200:
        print("[*] Failed logging in :: Please check email and password.")
        sys.exit()

    access_token = req.content.decode('utf-8')['token']

    client_data = load_client()
    product_data = load_product(exclude=['관세', '부가세', '통관료', '취급수수료'])


def load_product(exclude: list = []) -> list:
    # 데이터 로드:
    sheet = load_workbook(os.path.expanduser(PRODUCT_BOOK_PATH)).active
    book = Cells(sheet)
    product_data = []
    for row in range(1, book.max_row+1):
        product = {}
        product['name'] = book.get('B', row)
        product['specification'] = book.get('C', row)
        product['unit'] = book.get('D', row)
        product['price_a'] = book.get('F', row)
        product['price_b'] = book.get('G', row)
        product['price_c'] = book.get('H', row)
        product['price_d'] = book.get('I', row)
        product['price_e'] = book.get('J', row)
        product['essential_stock'] = book.get('N', row)
        product['memo'] = book.get('P', row)

        if product['name'] not in exclude:
            product_data.append(product)

    return product_data


def load_client(exclude: list = []) -> list:
    # 데이터 로드:
    sheet = load_workbook(os.path.expanduser(CLIENT_BOOK_PATH)).active
    book = Cells(sheet)
    client_data = []
    for row in range(1, book.max_row+1):
        client = {}
        client['name'] = book.get('B', row)
        client['crn'] = book.get('E', row)
        client['business_type'] = book.get('J', row)
        client['business_item'] = book.get('K', row)
        client['representative'] = book.get('D', row)
        client['poc1'] = book.get('Q', row)
        client['poc2'] = book.get('R', row)
        client['fax'] = book.get('S', row)
        client['worker'] = book.get('T', row)
        client['worker_email'] = book.get('V', row)
        client['worker_poc'] = book.get('U', row)
        client['default_price_type'] = 'A'
        client['postcode'] = book.get('G', row)
        client['address'] = str(book.get('H', row)) + str(book.get('I', row))
        client['memo'] = book.get('Z', row)
        client['staff_id'] = None
        client_data.append(client)

    # 데이터 가공:
    def sanitize(data):
        return data.replace('-', '').replace(' ', '').replace('.', '')
    for client in client_data:
        client['fax'] = sanitize(client['fax'])
        client['crn'] = sanitize(client['crn'])
        client['poc1'] = sanitize(client['poc1'])
        client['poc2'] = sanitize(client['poc2'])
        client['worker_poc'] = sanitize(client['worker_poc'])

    return client_data


if __name__ == "__main__":
    main()
